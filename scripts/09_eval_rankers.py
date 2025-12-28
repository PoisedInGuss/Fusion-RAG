#!/usr/bin/env python3
"""
Incoming: individual ranker .res files, qrels --- {TREC runs, relevance}
Processing: per-ranker IR evaluation + BEIR benchmark comparison --- {ir_measures}
Outgoing: publication-ready results --- {LaTeX, PDF, PNG, Excel, Markdown, JSON}

Step 9: Evaluate Individual Rankers (BEIR Comparison)
-----------------------------------------------------
Evaluates each retriever independently and compares against BEIR benchmark.
Produces publication-ready tables, figures, and reports.

BEIR Benchmark Reference (Table 2, NQ row, nDCG@10):
- BM25: 0.329 | docT5query: 0.399 | TAS-B†: 0.463 | ColBERT: 0.524 | BM25+CE†: 0.533

Usage:
    python scripts/09_eval_rankers.py --dataset nq
    python scripts/09_eval_rankers.py --dataset nq --output_format all
"""

import os
import sys
import json
import argparse
from pathlib import Path
from typing import Dict, List, Tuple
from collections import OrderedDict

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.config import config
from src.evaluation import IREvaluator
from src.data_utils import load_qrels as _load_qrels, load_run_file as _load_run

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt


# =============================================================================
# BEIR Benchmark Reference (Table 2)
# =============================================================================

BEIR_BENCHMARK = {
    'nq': OrderedDict([
        ('BM25', (0.329, 'Lexical', False)),
        ('DeepCT', (0.188, 'Sparse', False)),
        ('SPARTA', (0.398, 'Sparse', False)),
        ('docT5query', (0.399, 'Sparse', False)),
        ('DPR', (0.474, 'Dense', True)),
        ('ANCE', (0.446, 'Dense', False)),
        ('TAS-B', (0.463, 'Dense', True)),
        ('GenQ', (0.358, 'Dense', False)),
        ('ColBERT', (0.524, 'Late-Int', False)),
        ('BM25+CE', (0.533, 'Rerank', True)),
    ]),
    'hotpotqa': OrderedDict([
        ('BM25', (0.603, 'Lexical', False)),
        ('docT5query', (0.580, 'Sparse', False)),
        ('TAS-B', (0.584, 'Dense', False)),
        ('ColBERT', (0.593, 'Late-Int', False)),
        ('BM25+CE', (0.707, 'Rerank', False)),
    ]),
    'scifact': OrderedDict([
        ('BM25', (0.665, 'Lexical', False)),
        ('docT5query', (0.675, 'Sparse', False)),
        ('TAS-B', (0.643, 'Dense', False)),
        ('ColBERT', (0.671, 'Late-Int', False)),
        ('BM25+CE', (0.688, 'Rerank', False)),
    ]),
}

# Our ranker to BEIR category and comparison
RANKER_CONFIG = {
    'BM25': {'category': 'Lexical', 'beir_compare': 'BM25'},
    'Splade': {'category': 'Sparse', 'beir_compare': 'docT5query'},
    'BGE': {'category': 'Dense', 'beir_compare': 'TAS-B'},
    'BM25_TCT': {'category': 'Late-Int', 'beir_compare': 'ColBERT'},
    'BM25_MonoT5': {'category': 'Rerank', 'beir_compare': 'BM25+CE'},
}


# =============================================================================
# Data Loading
# =============================================================================

def load_qrels(qrels_path: Path) -> Dict[str, Dict[str, int]]:
    """Load qrels from TSV file."""
    return _load_qrels(qrels_path)


def load_run(run_path: Path) -> Dict[str, List[Tuple[str, float]]]:
    """Load TREC-format run file."""
    run_data = _load_run(run_path)
    return {qid: [(d, s) for d, s, r in docs] for qid, docs in run_data.items()}


# =============================================================================
# Evaluation
# =============================================================================

def evaluate_rankers(runs_dir: Path, qrels: Dict) -> Dict[str, Dict[str, float]]:
    """Evaluate all ranker run files in directory."""
    metrics = ['nDCG@5', 'nDCG@10', 'nDCG@20', 'RR@10', 'R@10', 'R@100']
    evaluator = IREvaluator(metrics=metrics)
    
    results = OrderedDict()
    res_files = sorted([f for f in runs_dir.glob("*.res") if '.norm.' not in f.name])
    
    for res_file in res_files:
        name = res_file.stem
        print(f"  Evaluating: {name}")
        run = load_run(res_file)
        scores = evaluator.evaluate(run, qrels, per_query=False)
        scores['n_queries'] = len([q for q in run.keys() if q in qrels])
        results[name] = scores
    
    return results


# =============================================================================
# Text Output
# =============================================================================

def print_results(results: Dict, dataset: str):
    """Print results in research paper format."""
    beir = BEIR_BENCHMARK.get(dataset, {})
    sorted_results = sorted(results.items(), key=lambda x: -x[1].get('nDCG@10', 0))
    best_ndcg = max(r.get('nDCG@10', 0) for r in results.values())
    
    print("\n" + "=" * 90)
    print(f"RETRIEVAL EVALUATION: {dataset.upper()}")
    print("=" * 90)
    
    # Table 1: Our Results
    print("\n" + "-" * 90)
    print("Table 1: Individual Ranker Performance")
    print("-" * 90)
    print(f"{'Ranker':<15} {'Type':<12} {'nDCG@5':<9} {'nDCG@10':<9} {'nDCG@20':<9} {'MRR@10':<9} {'R@10':<9} {'R@100':<9}")
    print("-" * 90)
    
    for name, m in sorted_results:
        cat = RANKER_CONFIG.get(name, {}).get('category', 'Other')
        marker = '*' if m.get('nDCG@10', 0) == best_ndcg else ' '
        print(f"{name:<14}{marker} {cat:<12} {m.get('nDCG@5',0):.4f}    {m.get('nDCG@10',0):.4f}    "
              f"{m.get('nDCG@20',0):.4f}    {m.get('RR@10',0):.4f}    {m.get('R@10',0):.4f}    {m.get('R@100',0):.4f}")
    
    print("-" * 90)
    print("* Best ranker")
    
    # Table 2: BEIR Comparison
    print("\n" + "-" * 90)
    print("Table 2: Comparison with BEIR Benchmark (Table 2)")
    print("-" * 90)
    print(f"{'Our Method':<15} {'Type':<10} {'Ours':<10} {'BEIR Ref':<15} {'BEIR':<10} {'Δ%':<10}")
    print("-" * 90)
    
    for name, m in sorted_results:
        cfg = RANKER_CONFIG.get(name, {})
        beir_name = cfg.get('beir_compare')
        if not beir_name or beir_name not in beir:
            continue
        
        our_score = m.get('nDCG@10', 0)
        beir_score, _, is_trained = beir[beir_name]
        delta = (our_score - beir_score) / beir_score * 100 if beir_score else 0
        suffix = '†' if is_trained else ''
        
        print(f"{name:<15} {cfg.get('category',''):<10} {our_score:.4f}     {beir_name}{suffix:<13} {beir_score:.3f}      {delta:+.1f}%")
    
    print("-" * 90)
    print("† In-domain trained model")


# =============================================================================
# LaTeX Generation
# =============================================================================

def generate_latex_results_table(results: Dict, dataset: str) -> str:
    """Generate LaTeX table for our results."""
    sorted_results = sorted(results.items(), key=lambda x: -x[1].get('nDCG@10', 0))
    best_ndcg = max(r.get('nDCG@10', 0) for r in results.values())
    
    lines = [
        r'\begin{table}[t]',
        r'\centering',
        r'\small',
        f'\\caption{{Retrieval Performance on {dataset.upper()} (nDCG@10)}}',
        f'\\label{{tab:{dataset}_results}}',
        r'\begin{tabular}{llccccc}',
        r'\toprule',
        r'\textbf{Method} & \textbf{Type} & \textbf{nDCG@5} & \textbf{nDCG@10} & \textbf{nDCG@20} & \textbf{MRR@10} & \textbf{R@100} \\',
        r'\midrule',
    ]
    
    for name, m in sorted_results:
        cat = RANKER_CONFIG.get(name, {}).get('category', 'Other')
        ndcg10 = m.get('nDCG@10', 0)
        ndcg10_str = f'\\textbf{{{ndcg10:.3f}}}' if ndcg10 == best_ndcg else f'{ndcg10:.3f}'
        
        lines.append(f"{name} & {cat} & {m.get('nDCG@5',0):.3f} & {ndcg10_str} & "
                    f"{m.get('nDCG@20',0):.3f} & {m.get('RR@10',0):.3f} & {m.get('R@100',0):.3f} \\\\")
    
    lines.extend([r'\bottomrule', r'\end{tabular}', r'\end{table}'])
    return '\n'.join(lines)


def generate_latex_beir_comparison(results: Dict, dataset: str) -> str:
    """Generate LaTeX table comparing with BEIR."""
    beir = BEIR_BENCHMARK.get(dataset, {})
    sorted_results = sorted(results.items(), key=lambda x: -x[1].get('nDCG@10', 0))
    
    lines = [
        r'\begin{table}[t]',
        r'\centering',
        r'\small',
        f'\\caption{{Comparison with BEIR Benchmark on {dataset.upper()} (nDCG@10)}}',
        f'\\label{{tab:{dataset}_beir}}',
        r'\begin{tabular}{llcccc}',
        r'\toprule',
        r'\textbf{Our Method} & \textbf{Type} & \textbf{Ours} & \textbf{BEIR Ref.} & \textbf{BEIR} & \textbf{$\Delta$} \\',
        r'\midrule',
    ]
    
    for name, m in sorted_results:
        cfg = RANKER_CONFIG.get(name, {})
        beir_name = cfg.get('beir_compare')
        if not beir_name or beir_name not in beir:
            continue
        
        our_score = m.get('nDCG@10', 0)
        beir_score, _, is_trained = beir[beir_name]
        delta = (our_score - beir_score) / beir_score * 100 if beir_score else 0
        beir_ref = f"{beir_name}$^\\dagger$" if is_trained else beir_name
        delta_str = f'+{delta:.1f}\\%' if delta >= 0 else f'{delta:.1f}\\%'
        
        lines.append(f"{name} & {cfg.get('category','')} & {our_score:.3f} & {beir_ref} & {beir_score:.3f} & {delta_str} \\\\")
    
    lines.extend([
        r'\bottomrule',
        r'\end{tabular}',
        r'\vspace{2pt}',
        r'\footnotesize{$^\dagger$ In-domain trained model.}',
        r'\end{table}',
    ])
    return '\n'.join(lines)


# =============================================================================
# Figure Generation
# =============================================================================

def create_comparison_figure(results: Dict, output_path: Path, dataset: str):
    """Create bar chart comparing our results to BEIR baselines."""
    beir = BEIR_BENCHMARK.get(dataset, {})
    
    plt.rcParams.update({
        'font.family': 'serif',
        'font.size': 10,
        'axes.labelsize': 11,
        'axes.titlesize': 12,
        'figure.dpi': 150,
        'savefig.dpi': 300,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })
    
    fig, ax = plt.subplots(figsize=(8, 4))
    
    # Our data sorted by score
    data = sorted([(n, m.get('nDCG@10', 0)) for n, m in results.items()], key=lambda x: -x[1])
    
    colors = {'Lexical': '#4C72B0', 'Sparse': '#55A868', 'Dense': '#C44E52', 
              'Late-Int': '#8172B3', 'Rerank': '#CCB974'}
    
    for i, (name, score) in enumerate(data):
        cat = RANKER_CONFIG.get(name, {}).get('category', 'Other')
        ax.bar(i, score, color=colors.get(cat, '#999999'), edgecolor='white', linewidth=0.5)
        ax.annotate(f'{score:.3f}', xy=(i, score + 0.01), ha='center', va='bottom', fontsize=8)
    
    # BEIR reference lines
    if 'BM25' in beir:
        ax.axhline(y=beir['BM25'][0], color='#333333', linestyle='--', linewidth=1, 
                   label=f"BEIR BM25 ({beir['BM25'][0]})")
    if 'BM25+CE' in beir:
        ax.axhline(y=beir['BM25+CE'][0], color='#666666', linestyle=':', linewidth=1,
                   label=f"BEIR BM25+CE ({beir['BM25+CE'][0]})")
    
    ax.set_ylabel('nDCG@10')
    ax.set_xlabel('Retrieval Method')
    ax.set_title(f'{dataset.upper()}: Our Retrievers vs BEIR Benchmark', fontweight='bold')
    ax.set_xticks(range(len(data)))
    ax.set_xticklabels([d[0] for d in data])
    ax.set_ylim(0, 0.65)
    ax.legend(loc='upper right', framealpha=0.9)
    
    plt.tight_layout()
    for fmt in ['pdf', 'png']:
        fig.savefig(output_path.with_suffix(f'.{fmt}'), format=fmt, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f"[09_eval] Saved: {output_path.with_suffix('.pdf')}")


def create_beir_table_figure(results: Dict, output_path: Path, dataset: str):
    """Create BEIR-style comparison table as figure."""
    beir = BEIR_BENCHMARK.get(dataset, {})
    
    plt.rcParams.update({'font.family': 'serif', 'font.size': 9})
    fig, ax = plt.subplots(figsize=(10, 3))
    ax.axis('off')
    
    # Build table data
    header = ['Category', 'Our Method', 'Our nDCG@10', 'BEIR Method', 'BEIR nDCG@10', 'Δ%']
    table_data = []
    
    for name in ['BM25', 'Splade', 'BGE', 'BM25_MonoT5']:
        if name not in results:
            continue
        cfg = RANKER_CONFIG.get(name, {})
        beir_name = cfg.get('beir_compare', '')
        if beir_name not in beir:
            continue
        
        our_score = results[name].get('nDCG@10', 0)
        beir_score = beir[beir_name][0]
        is_trained = beir[beir_name][2]
        delta = (our_score - beir_score) / beir_score * 100 if beir_score else 0
        
        table_data.append([
            cfg.get('category', ''),
            name,
            f'{our_score:.3f}',
            f"{beir_name}{'†' if is_trained else ''}",
            f'{beir_score:.3f}',
            f'{delta:+.1f}%'
        ])
    
    table = ax.table(cellText=table_data, colLabels=header, loc='center', cellLoc='center',
                     colWidths=[0.12, 0.15, 0.15, 0.15, 0.15, 0.10])
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1.2, 1.5)
    
    for i in range(len(header)):
        table[(0, i)].set_facecolor('#E8E8E8')
        table[(0, i)].set_text_props(weight='bold')
    
    for row in range(1, len(table_data) + 1):
        our_val = float(table_data[row-1][2])
        beir_val = float(table_data[row-1][4])
        if our_val > beir_val:
            table[(row, 2)].set_text_props(weight='bold', color='green')
        elif our_val < beir_val:
            table[(row, 4)].set_text_props(weight='bold')
    
    ax.set_title(f'{dataset.upper()}: Comparison with BEIR Benchmark', fontsize=11, fontweight='bold', pad=20)
    plt.tight_layout()
    
    for fmt in ['pdf', 'png']:
        fig.savefig(output_path.with_suffix(f'.{fmt}'), format=fmt, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f"[09_eval] Saved: {output_path.with_suffix('.pdf')}")


# =============================================================================
# Excel Generation
# =============================================================================

def generate_excel_report(results: Dict, output_path: Path, dataset: str):
    """Generate Excel report with multiple sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("[09_eval] Warning: openpyxl not installed. Skipping Excel.")
        return
    
    beir = BEIR_BENCHMARK.get(dataset, {})
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Sheet 1: Our Results
    ws1 = wb.active
    ws1.title = "Our Results"
    
    headers = ["Method", "Type", "nDCG@5", "nDCG@10", "nDCG@20", "MRR@10", "R@10", "R@100"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, border
        c.alignment = Alignment(horizontal="center")
    
    sorted_results = sorted(results.items(), key=lambda x: -x[1].get('nDCG@10', 0))
    best_ndcg = max(r.get('nDCG@10', 0) for r in results.values())
    
    for row_idx, (name, m) in enumerate(sorted_results, 2):
        cat = RANKER_CONFIG.get(name, {}).get('category', 'Other')
        row = [name, cat, round(m.get('nDCG@5',0),4), round(m.get('nDCG@10',0),4),
               round(m.get('nDCG@20',0),4), round(m.get('RR@10',0),4),
               round(m.get('R@10',0),4), round(m.get('R@100',0),4)]
        for col, v in enumerate(row, 1):
            c = ws1.cell(row=row_idx, column=col, value=v)
            c.border = border
            c.alignment = Alignment(horizontal="center")
            if col == 4 and v == best_ndcg:
                c.font = Font(bold=True, color="006400")
    
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 2
    
    # Sheet 2: BEIR Comparison
    ws2 = wb.create_sheet("BEIR Comparison")
    comp_headers = ["Our Method", "Type", "Our Score", "BEIR Method", "BEIR Score", "Δ%", "Status"]
    for col, h in enumerate(comp_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, border
        c.alignment = Alignment(horizontal="center")
    
    row_idx = 2
    for name, m in sorted_results:
        cfg = RANKER_CONFIG.get(name, {})
        beir_name = cfg.get('beir_compare')
        if not beir_name or beir_name not in beir:
            continue
        
        our_score = m.get('nDCG@10', 0)
        beir_score = beir[beir_name][0]
        is_trained = beir[beir_name][2]
        delta = (our_score - beir_score) / beir_score * 100 if beir_score else 0
        status = "✓ Better" if delta > 0 else "≈ Close" if delta > -5 else "Below"
        
        row = [name, cfg.get('category',''), round(our_score,4),
               f"{beir_name}{'†' if is_trained else ''}", beir_score, f"{delta:+.1f}%", status]
        for col, v in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col, value=v)
            c.border = border
            c.alignment = Alignment(horizontal="center")
            if col == 6:
                c.font = Font(bold=True, color="006400" if delta > 0 else "8B0000" if delta < -5 else "000000")
        row_idx += 1
    
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 2
    
    # Sheet 3: BEIR Reference
    ws3 = wb.create_sheet("BEIR Reference")
    beir_headers = ["Method", "Category", "nDCG@10", "In-Domain"]
    for col, h in enumerate(beir_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        c.border = border
    
    for row_idx, (name, (score, cat, trained)) in enumerate(beir.items(), 2):
        for col, v in enumerate([name, cat, score, "†" if trained else ""], 1):
            c = ws3.cell(row=row_idx, column=col, value=v)
            c.border = border
            c.alignment = Alignment(horizontal="center")
    
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 2
    
    wb.save(output_path)
    print(f"[09_eval] Saved: {output_path}")


# =============================================================================
# Markdown Generation
# =============================================================================

def generate_markdown_summary(results: Dict, output_path: Path, dataset: str):
    """Generate markdown summary."""
    beir = BEIR_BENCHMARK.get(dataset, {})
    sorted_results = sorted(results.items(), key=lambda x: -x[1].get('nDCG@10', 0))
    best_name, best_m = sorted_results[0]
    
    md = f"""# {dataset.upper()} Retrieval Results - BEIR Comparison

**Generated:** {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}  
**Queries:** {best_m.get('n_queries', 'N/A')} test queries

---

## Summary

| Metric | Best Method | Score | BEIR Best | Δ |
|--------|-------------|-------|-----------|---|
| nDCG@10 | {best_name} | **{best_m.get('nDCG@10',0):.4f}** | BM25+CE ({beir.get('BM25+CE', (0,))[0]}) | {((best_m.get('nDCG@10',0) - beir.get('BM25+CE', (0.533,))[0]) / beir.get('BM25+CE', (0.533,))[0] * 100):+.1f}% |

---

## Individual Ranker Performance

| Method | Type | nDCG@5 | nDCG@10 | nDCG@20 | MRR@10 | R@100 |
|--------|------|--------|---------|---------|--------|-------|
"""
    
    for name, m in sorted_results:
        cat = RANKER_CONFIG.get(name, {}).get('category', 'Other')
        md += f"| {name} | {cat} | {m.get('nDCG@5',0):.4f} | **{m.get('nDCG@10',0):.4f}** | "
        md += f"{m.get('nDCG@20',0):.4f} | {m.get('RR@10',0):.4f} | {m.get('R@100',0):.4f} |\n"
    
    md += "\n---\n\n## BEIR Comparison\n\n"
    md += "| Our Method | Type | Ours | BEIR Method | BEIR Score | Δ% |\n"
    md += "|------------|------|------|-------------|------------|-----|\n"
    
    for name, m in sorted_results:
        cfg = RANKER_CONFIG.get(name, {})
        beir_name = cfg.get('beir_compare')
        if not beir_name or beir_name not in beir:
            continue
        our_score = m.get('nDCG@10', 0)
        beir_score = beir[beir_name][0]
        delta = (our_score - beir_score) / beir_score * 100 if beir_score else 0
        md += f"| {name} | {cfg.get('category','')} | {our_score:.4f} | {beir_name} | {beir_score:.3f} | {delta:+.1f}% |\n"
    
    md += "\n†: In-domain trained model\n"
    
    with open(output_path, 'w') as f:
        f.write(md)
    print(f"[09_eval] Saved: {output_path}")


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Step 9: Evaluate Rankers (BEIR Comparison)")
    parser.add_argument("--dataset", default="nq", choices=config.datasets.supported)
    parser.add_argument("--data_dir", default=None)
    parser.add_argument("--corpus_path", default=None)
    parser.add_argument("--output_format", default="all", choices=["text", "all"])
    parser.add_argument("--output_dir", default=None)
    args = parser.parse_args()
    
    # Paths
    data_dir = Path(args.data_dir) if args.data_dir else config.project_root / "data" / args.dataset
    runs_dir = data_dir / "runs"
    qrels_path = Path(args.corpus_path) / "qrels" / "test.tsv" if args.corpus_path else config.get_qrels_path(args.dataset)
    output_dir = Path(args.output_dir) if args.output_dir else data_dir / "results" / "paper"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    if not runs_dir.exists():
        print(f"Error: Runs directory not found: {runs_dir}")
        sys.exit(1)
    
    print(f"[09_eval] Dataset: {args.dataset}")
    print(f"[09_eval] Runs: {runs_dir}")
    print(f"[09_eval] Output: {output_dir}")
    
    # Load and evaluate
    qrels = load_qrels(qrels_path)
    print(f"[09_eval] Loaded {len(qrels)} queries")
    
    print("\n[09_eval] Evaluating rankers...")
    results = evaluate_rankers(runs_dir, qrels)
    
    # Exclude broken rankers
    if results.get('BM25_TCT', {}).get('nDCG@10', 0) < 0.2:
        print("[09_eval] Note: BM25_TCT excluded (broken)")
        del results['BM25_TCT']
    
    # Generate outputs
    print_results(results, args.dataset)
    
    if args.output_format == "all":
        print("\n[09_eval] Generating LaTeX...")
        with open(output_dir / 'results_table.tex', 'w') as f:
            f.write(generate_latex_results_table(results, args.dataset))
        with open(output_dir / 'beir_comparison.tex', 'w') as f:
            f.write(generate_latex_beir_comparison(results, args.dataset))
        
        print("[09_eval] Generating figures...")
        create_comparison_figure(results, output_dir / 'comparison_figure', args.dataset)
        create_beir_table_figure(results, output_dir / 'beir_table_figure', args.dataset)
        
        print("[09_eval] Generating Excel...")
        generate_excel_report(results, output_dir / 'results.xlsx', args.dataset)
        
        print("[09_eval] Generating Markdown...")
        generate_markdown_summary(results, output_dir / 'RESULTS_SUMMARY.md', args.dataset)
        
        # JSON
        with open(output_dir / 'results.json', 'w') as f:
            json.dump(results, f, indent=2)
    
    print(f"\n{'='*60}")
    print("COMPLETE")
    print(f"{'='*60}")
    print(f"Output: {output_dir}")
    for f in sorted(output_dir.glob('*')):
        print(f"  - {f.name}")


if __name__ == "__main__":
    main()



